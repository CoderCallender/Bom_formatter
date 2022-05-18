
from msilib.schema import Font
import os
from pickle import FALSE
import shutil
from tkinter.font import BOLD


#FUNCTIONS
def check_header_valid(string):

    wanted_cols = ["Qty","Value","Device","Package","Parts","Description","MF","MPN","OC_MOUSER","OC_FARNELL","OC_DIGIKEY","ALT1","ALT_1"]

    for x in wanted_cols:
        if x in string:
            #header/title is to stay
            return True
    
    #unwanted header
    return False


source_file = r'C:\Users\callende\Desktop\Python Scripts\RAW_BOM.xlsx'
working_file = r'C:\Users\callende\Desktop\Python Scripts\WORKING_BOM.xlsx'



#check if there is a file already - delete if so

if os.path.exists(working_file):
    os.remove(working_file)
else:
    print("No file to delete.")

# make a copy of our file
shutil.copyfile(source_file, working_file)

#get the data from the new file
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
workbook = load_workbook(working_file)

#save active spreadsheet
spreadsheet = workbook.active

#get all the titles and delete ones we don't like
col_index = 0
col_to_delete = [] #create empty list

print("deleted coloums:")
for value in spreadsheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=50, values_only=True):
   
    col_index = col_index + 1
 
   # print("value = ", value, col_index)

    if check_header_valid(value) == False:
        print(value)
        col_to_delete.append(col_index)

        
  #  else:
   #     col_index = col_index + 1

#now we have the list, delete and adjust appropriately 

for x in range(0 ,len(col_to_delete)):
 #   print("number", x)
 #   print("col to delete abs",col_to_delete[x])
 #   print("col to delete ajust",(col_to_delete[x]-x))
    spreadsheet.delete_cols((col_to_delete[x]-x),1)

#find out where the end of the data is in col
col_index_max = 0
for col in spreadsheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=50):
    for cell in col:
        if cell.value:
            col_index_max = col_index_max + 1
      #  else:
         #   print("max col =",col_index_max)
     #   break
#print("max col =",col_index_max)
#max_col_letter = get_column_letter(col_index_max)
#print("max col letter =",max_col_letter)
    

#find out where the end of the data is in row
row_index_max = 0
for row in spreadsheet.iter_rows(min_row=1, max_row=50, min_col=1, max_col=1):
    for cell in row:
        if cell.value:
            row_index_max = row_index_max + 1
   
#print("max row =",row_index_max)

#format
from openpyxl.styles import Font, PatternFill, Border, Side

top_bottom_thick_border = Border(left=Side(style='none'),
                           right=Side(style='none'),
                           top=Side(style='thick'),
                           bottom=Side(style='thick'), )

top_bottom_left_thick_border = Border(left=Side(style='thick'),
                           right=Side(style='none'),
                           top=Side(style='thick'),
                           bottom=Side(style='thick'), )

top_bottom_right_thick_border = Border(left=Side(style='none'),
                           right=Side(style='thick'),
                           top=Side(style='thick'),
                           bottom=Side(style='thick'), )

all_thick_border = Border(left=Side(style='thick'),
                           right=Side(style='thick'),
                           top=Side(style='thick'),
                           bottom=Side(style='thick'), )

for x in range(1, col_index_max + 1):
        spreadsheet.cell(1,x).font = Font(bold = True)
        spreadsheet.cell(1,x).fill = PatternFill(start_color='00FFCC99',end_color='00FFCC99',fill_type='solid')
    #    if x == 1:
    #    spreadsheet.cell(1,x).border = all_thick_border
    #    elif x == col_index_max:
    #        spreadsheet.cell(1,x).border = top_bottom_right_right_border
    #    else:
    #        spreadsheet.cell(1,x).border = top_bottom_thick_border
        



#check for duplicates
print("value duplicates")
for x in range(1, row_index_max + 1):
    if spreadsheet.cell(x,2).value:
        master_val = spreadsheet.cell(x,2).value
        master_val = str (master_val)   #typecast to string
#master_val = master_val.lower()
        master_pos = x  #save position

    for y in range(1, row_index_max + 1):
        if spreadsheet.cell(y,2).value:
            check_val = spreadsheet.cell(y,2).value
            check_val = str(check_val)  #cast to string
            if master_val.lower() in check_val.lower():
                if x != y:
                    spreadsheet.cell(1,col_index_max+1).value = "VALUE DUPLICATE ON ROW"      #set red title for the duplicates
                    spreadsheet.cell(1,col_index_max+1).font = Font(color = "00FF0000")
                    spreadsheet.cell(x,col_index_max+1).value = y     #save duplicate location on spreadsheet
                    spreadsheet.cell(x,col_index_max+1).font = Font(color = "00FF0000")
                 #   print("warning - duplicate found!")
                 #   print("master",master_val)
                 #   print("check",check_val)
                 #   print("x",x)
                 #   print("y",y)



workbook.save(working_file)